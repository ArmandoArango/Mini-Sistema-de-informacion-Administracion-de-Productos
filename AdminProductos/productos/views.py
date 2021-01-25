
from __future__ import unicode_literals

from django.shortcuts import render
from django.views.generic.base import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.contrib import messages

from AdminProductosApp.models import Producto
from AdminProductosApp.models import Usuarios
from AdminProductosApp.forms import ProductoForm
from AdminProductosApp.forms import ProductForm2
from openpyxl import  Workbook
from django.http.response import HttpResponse

# Create your views here.


class ListadoProducto(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'productos/listar_producto.html'

    def get(self, request):
        '''
        lista de los productos
        :param request:
        :return: retorna la lisata de todo los producto
        '''
        lista_productos = Producto.objects.filter()
        return render(request, self.template_name, {'productos': lista_productos })


class VisualizarProducto(LoginRequiredMixin, View):
    login_url = '/'
    form_class = ProductForm2
    template_name = 'productos/visualizar_producto.html'

    def get(self, request, id_producto):
        '''

        :param request:
        :param id_producto: parametro de llave primaria
        :return: devuelve el producto con todos sus datos
        '''

        try:
            datos_cliente = Producto.objects.get(id=id_producto)
            form = self.form_class(instance=datos_cliente)
            return render(request, self.template_name, {'form': form})

        except Producto.DoesNotExist:
            list = ListadoProducto()
            return list.get(request)


class AdicionarProducto(LoginRequiredMixin, View):
    login_url = '/'
    form_class = ProductoForm
    template_name = 'productos/adicionar_producto.html'

    def get(self, request):
        '''

        :param request:
        :return: retorna el formulario de adiccionar
        '''
        try:
            form = self.form_class()
            return render(request,
                          self.template_name,
                          {'form': form,})

        except Producto.DoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        '''

        :param request:
        :return: retorna un objecto producto ya creado
        '''
        user = Usuarios.objects.get(usuid=request.user.pk)
        try:
            nombre = request.POST.get('nombre', None)
            precio = request.POST.get('precio', None)
            informacion = request.POST.get('informacion', None)
            fecha_adquision = request.POST.get('fecha_adquision', None)
            image = request.POST.get('image', None)

            producto = Producto(nombre=nombre,
                                precio=precio,
                                informacion=informacion,
                                fecha_adquision=fecha_adquision,
                                image=image,
                                usuario=user)

            producto.save()
            messages.add_message(request, messages.INFO, 'El producto se adiciono correctamente')

            list= ListadoProducto()
            return list.get(request)

        except Producto.DoesNotExist:
            return render(request, "pages-404.html")


class ModificarProducto(LoginRequiredMixin, View):
    login_url = '/'
    form_class = ProductForm2
    template_name = 'productos/modificar_producto.html'

    def get(self, request, id_producto):
        '''

        :param request:
        :param id_producto: el identificador de cada producto
        :return: retorna el objecto a modificar
        '''
        try:
            producto = Producto.objects.get(id=id_producto)
            form = self.form_class(instance=producto)
            return render(request,
                          self.template_name,
                          {'form': form})

        except Producto.DoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, id_producto):
        '''

        :param request:
        :param id_producto: el identificador de cada producto
        :return: retorna el objecto modificado
        '''
        try:

            datos_produto = Producto.objects.get(id=id_producto)
            form = self.form_class(request.POST, instance=datos_produto)

            if form.is_valid:
                form.save()
                messages.add_message(request, messages.INFO, "EL producto se modico coorectamente")
            else:
                messages.add_message(request, messages.INFO, "No se pudo modificar el Producto")

        except Producto.DoesNotExist:
            messages.add_message(request, messages.ERROR, "ERROR INTENTE NUEVAMENTE")

        list = ListadoProducto()
        return list.get(request)


class EliminarProducto(LoginRequiredMixin, View):
    login_url = '/'
    form_class = ProductForm2
    template_name = 'productos/eliminar_producto.html'

    def get(self, request, id_producto):
        '''

        :param request:
        :param id_producto: el identificador de cada producto
        :return: el objecto a eliminar
        '''
        try:
            datos_producto = Producto.objects.get(id=id_producto)
            form = self.form_class(instance=datos_producto)
            return render(request, self.template_name, {'form': form})

        except Producto.DoesNotExist:
            list = ListadoProducto()
            return list.get(request)

    def post(self, request, id_producto,):
        '''

        :param request:
        :param id_producto: el identificador de cada producto
        :return: retorna un mensaje de confirmacion cuando se elimina el objecto
        '''
        try:
            datos_producto = Producto.objects.get(id=id_producto)
            datos_producto.delete()
            messages.add_message(request, messages.INFO, "El producto se borró correctamente")

        except Producto.DoesNotExist:
            messages.add_message(request, messages.ERROR, "Error al borrar intente nuevamente")

        list = ListadoProducto()
        return list.get(request)


class ReporteExcel(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        '''

        :param request:
        :param args: argumentos
        :param kwargs: argumentos
        :return: una lista de productos existentes en la base de datos
        '''
        productos = Producto.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Reporte De Prductos'

        ws.merge_cells('A1:D1')

        ws['A2'] = 'Nombre'
        ws['B2'] = 'Precio'
        ws['C2'] = 'Informacion'
        ws['D2'] = 'Fecha de Adquisicion'

        cont = 3

        for producto in productos:
            ws.cell(row=cont, column=1).value = producto.nombre
            ws.cell(row=cont, column=2).value = producto.precio
            ws.cell(row=cont, column=3).value = producto.informacion
            ws.cell(row=cont, column=4).value = producto.fecha_adquision
            cont += 1

        nombre_archivo = 'ReporteExcel.xlsx'
        response = HttpResponse(content_type = 'application/ms-excel')
        content = 'attachment; filename = {0}'.format(nombre_archivo)
        response['Content-Diposition'] = content
        wb.save(response)
        return response